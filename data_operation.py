"""
@Description: 
@Author: Shen Du
@Date: 2024-04-17 18:48:56
@LastEditTime: 2024-04-22 10:41:34
@LastEditors: Ambrose
"""

import os

from PIL import Image
import numpy as np
import jieba
import openpyxl
from wordcloud import WordCloud

from setting import setting as st


class data_operate(object):
    def __init__(self):
        self.file_path = st.file_info.keys()
        self.text = ""
        self.words = []
        self.dict = {}

    # 数据读入
    def data_read(self, file_name):
        with open(file_name, "r", encoding="utf-8") as f:
            self.text = f.read()
        f.close()

    # 数据分割与数据清洗
    def data_split(self):
        self.dict = {}
        # 特殊词标注
        if len(st.special_word_list) != 0:
            for word in st.special_word_list:
                jieba.suggest_freq(word, tune=True)
        # 从文件导入停用词表
        stop_dic = open(st.stop_words_path, "rb")
        stop_word_content = stop_dic.read()
        stop_word_list = stop_word_content.splitlines()
        stop_dic.close()
        segs = jieba.cut(self.text)
        # 文本清洗
        for seg in segs:
            if seg not in stop_word_list and seg != " " and len(seg) != 1:
                self.words.append(seg.replace(" ", " "))

        for word in self.words:
            if word in self.dict:
                self.dict[word] += 1
            else:
                self.dict[word] = 1
        self.dict = sorted(self.dict.items(), key=lambda x: x[1], reverse=True)
        self.dict = dict(self.dict)

    # 以txt方式数据存储
    def data_storage_txt(self, file_name):
        if os.path.exists("./output/txt/{}".format(file_name)):
            os.remove("./output/txt/{}".format(file_name))
        with open("./output/txt/{}".format(file_name), "w", encoding="utf-8") as f:
            for word, count in self.dict.items():
                f.write(f"{word}: {count}")
                f.write("\n")
        f.close()

    # 以csv方式数据存储
    def data_storage_csv(self, file_name):
        if os.path.exists("./output/csv/{}.csv".format(file_name[:4])):
            os.remove("./output/csv/{}.csv".format(file_name[:4]))
        with open(
            "./output/csv/{}.csv".format(file_name[:4]), "w", encoding="utf-8"
        ) as f:
            for word, count in self.dict.items():
                f.write(f"{word},{count}")
                f.write("\n")
        f.close()

    # 以excel方式数据存储
    def data_storage_xlsx(self, file_name):
        if os.path.exists("./output/excel/{}.xlsx".format(file_name[:4])):
            os.remove("./output/excel/{}.xlsx".format(file_name[:4]))
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet["A1"] = "词"
        sheet["B1"] = "出现数量"
        row = 2
        for word, count in self.dict.items():
            sheet.cell(row, column=1, value=word)
            sheet.cell(row, column=2, value=count)
            row += 1
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length + 2
        workbook.save("./output/excel/{}.xlsx".format(file_name[:4]))

    # 生成词云
    def word_cloud_generate(self, file_name):
        shape_mask = np.array(Image.open(st.background_img))
        text = " ".join(self.words)
        word_cloud = WordCloud(
            background_color="white",
            font_path=st.font_path,
            mask=shape_mask,
            width=1000,
            height=700,
            max_words=50
        ).generate(text)
        word_cloud.to_file("./output/word_cloud/wordCloud{}.png".format(file_name[:4]))
       
    # 数据操作入口
    def run(self, save):
        for file_name in self.file_path:
            self.data_read("./resource/{}".format(file_name))
            self.data_split()

            if save == "txt":
                self.data_storage_txt(file_name)
            elif save == "csv":
                self.data_storage_csv(file_name)
            elif save == "xlsx":
                self.data_storage_xlsx(file_name)
            elif save == "all":
                self.data_storage_txt(file_name)
                self.data_storage_csv(file_name)
                self.data_storage_xlsx(file_name)

            self.word_cloud_generate(file_name)


if __name__ == "__main__":
    operate = data_operate()
    operate.run(save="all")
